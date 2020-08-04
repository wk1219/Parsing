import urllib.request as req
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
#-*- coding: utf-8 -*-

URL = input("Input URL : ")
file_name = input("Input file path : ")
r = requests.get(URL)
# print(r.status_code)
# print(r.encoding)

res = req.urlopen(URL).read()
content = res.decode('utf-8', 'replace')                    # Data decoding : UTF-8
soup = BeautifulSoup(content, 'html.parser')

# Data Parsing & Processing
app_name = soup.find_all('div', {'class':'WsMG1c nnK0zc'})                # App Name property
app_maker = soup.find_all('div', {'class':'b8cIId ReQCgd KoLSrc'})        # App Maker property
# print(len(app_name))
# print(len(app_maker))

name_list = []
maker_list = []

for i in range(0, len(app_name)):
    name_list.append(app_name[i].get_text("", strip=True))
    maker_list.append(app_maker[i].get_text("", strip=True))
    print(name_list[i] + '%' + maker_list[i])
print("총 개수 : %d" % len(app_name))

load_wb = load_workbook(file_name)
load_ws = load_wb['시트1']

for i in range(0, len(name_list)):
    if load_ws.cell(i + 1, 3).value == None and load_ws.cell(i + 1, 4).value == None:
        load_ws.cell(i + 1, 3, maker_list[i])
        load_ws.cell(i + 1, 4, name_list[i])
    else:
        pass
load_wb.save(file_name)